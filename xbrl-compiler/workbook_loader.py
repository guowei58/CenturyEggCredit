"""Load Excel workbooks, detect statement tabs, extract facts, classify 10-K."""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook as _xl_open

from period_parser import Period, parse_period

logger = logging.getLogger(__name__)

# ── tab-name → statement_type ──────────────────────────────────────────────
_EXACT_TAB: dict[str, str] = {
    "income statement": "income_statement",
    "balance sheet": "balance_sheet",
    "cash flow": "cash_flow",
    "cash flow statement": "cash_flow",
    "cash flows": "cash_flow",
}

_KEYWORD_RULES: list[tuple[str, str]] = [
    ("statement of operations", "income_statement"),
    ("statements of operations", "income_statement"),
    ("statements of income", "income_statement"),
    ("consolidated income", "income_statement"),
    ("statement of cash flows", "cash_flow"),
    ("statements of cash flows", "cash_flow"),
    ("consolidated balance sheet", "balance_sheet"),
    ("financial position", "balance_sheet"),
]


def classify_tab(raw_name: str) -> str | None:
    """Map a worksheet name to a statement type, or None to skip."""
    norm = re.sub(r"[^a-z0-9\s]", "", raw_name.lower()).strip()
    if norm in _EXACT_TAB:
        return _EXACT_TAB[norm]
    for kw, st in _KEYWORD_RULES:
        if kw in norm:
            return st
    return None


# ── 10-K detection ─────────────────────────────────────────────────────────
_10K_RE = re.compile(r"10[_\-\s]?K", re.I)


def is_10k_filename(name: str) -> bool:
    return bool(_10K_RE.search(name))


# ── data classes ───────────────────────────────────────────────────────────
@dataclass
class FactRecord:
    statement_type: str
    concept: str
    line_label: str
    period: Period
    value: float | None
    source_file: str
    source_sheet: str
    source_column: str
    depth: int = 0


@dataclass
class SheetData:
    source_file: str
    source_sheet: str
    statement_type: str
    facts: list[FactRecord]
    row_order: list[str]          # concept order as found
    concept_to_line: dict[str, str] = field(default_factory=dict)
    concept_to_depth: dict[str, int] = field(default_factory=dict)


@dataclass
class WorkbookInfo:
    filename: str
    filepath: Path
    is_10k: bool
    latest_fy: int
    sheets: list[SheetData]


# ── header scanning ────────────────────────────────────────────────────────
def _find_header(ws: Any, max_rows: int = 10) -> tuple[int, dict[int, str]] | None:
    for ri in range(1, max_rows + 1):
        cells: dict[int, str] = {}
        for ci, cell in enumerate(ws[ri], 1):
            v = str(cell.value).strip() if cell.value is not None else ""
            if v:
                cells[ci] = v
        if any(v.lower() == "concept" for v in cells.values()) and any(
            parse_period(v) is not None for v in cells.values()
        ):
            return ri, cells
    return None


SPARSE_COLUMN_THRESHOLD = 0.50  # columns below 50% of max fill are dropped


def _filter_sparse_columns(
    ws: Any,
    header_row: int,
    pcols: list[tuple[int, Period]],
) -> list[tuple[int, Period]]:
    """
    Count non-empty numeric cells in each period column.
    Drop any column whose fill count is below SPARSE_COLUMN_THRESHOLD of the
    most-populated column.  This removes footnote / dimensional columns that
    carry only a handful of values.
    """
    fill_counts: dict[int, int] = {ci: 0 for ci, _ in pcols}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        for ci, _ in pcols:
            cell = row[ci - 1] if ci - 1 < len(row) else None
            raw = cell.value if cell else None
            if raw is None or raw == "" or raw == "\u2014":
                continue
            try:
                float(raw)
                fill_counts[ci] += 1
            except (ValueError, TypeError):
                continue

    if not fill_counts:
        return pcols

    max_fill = max(fill_counts.values())
    if max_fill == 0:
        return pcols

    cutoff = max_fill * SPARSE_COLUMN_THRESHOLD
    kept: list[tuple[int, Period]] = []
    for ci, period in pcols:
        cnt = fill_counts[ci]
        if cnt >= cutoff:
            kept.append((ci, period))
        else:
            logger.info("    Dropping sparse column '%s' (%d/%d = %.0f%% fill)",
                        period.column_label, cnt, max_fill,
                        100 * cnt / max_fill if max_fill else 0)

    return kept


# ── single workbook ────────────────────────────────────────────────────────
def load_workbook_data(filepath: str | Path) -> WorkbookInfo:
    fp = Path(filepath)
    fname = fp.name
    logger.info("Loading %s", fname)

    wb = _xl_open(str(fp), data_only=True)
    sheets: list[SheetData] = []
    max_fy = 0

    for tab_name in wb.sheetnames:
        stmt = classify_tab(tab_name)
        if stmt is None:
            continue

        ws = wb[tab_name]
        hdr = _find_header(ws)
        if hdr is None:
            logger.warning("  '%s': no header row", tab_name)
            continue

        hrow, hcells = hdr
        concept_col: int | None = None
        line_col: int | None = None
        depth_col: int | None = None
        pcols: list[tuple[int, Period]] = []

        for ci, val in hcells.items():
            low = val.lower().strip()
            if low == "concept":
                concept_col = ci
            elif low == "line":
                line_col = ci
            elif low == "depth":
                depth_col = ci
            else:
                p = parse_period(val)
                if p:
                    pcols.append((ci, p))

        if concept_col is None or not pcols:
            continue

        # ── sparse-column filter: drop period columns filled < 50% of the max ─
        pcols = _filter_sparse_columns(ws, hrow, pcols)
        if not pcols:
            logger.warning("  '%s': all period columns filtered as sparse", tab_name)
            continue

        facts: list[FactRecord] = []
        order: list[str] = []
        seen: set[str] = set()
        c2line: dict[str, str] = {}
        c2depth: dict[str, int] = {}

        for ri, row in enumerate(ws.iter_rows(min_row=hrow + 1, values_only=False), start=1):
            cc = row[concept_col - 1] if concept_col - 1 < len(row) else None
            concept = str(cc.value).strip() if cc and cc.value is not None else ""

            lc = row[line_col - 1] if line_col and line_col - 1 < len(row) else None
            line = str(lc.value).strip() if lc and lc.value is not None else ""

            depth = 0
            if depth_col:
                dc = row[depth_col - 1] if depth_col - 1 < len(row) else None
                try:
                    depth = int(dc.value) if dc and dc.value is not None else 0
                except (ValueError, TypeError):
                    pass

            if not concept and not line:
                continue

            # Line item with a face label but no QName in the Concept column
            # (common in some Excel exports) — still needs a stable key so the
            # row participates in row_order and the compiler pipeline.
            if not concept and line:
                concept = f"_:lineonly:{fname}|{tab_name}|R{hrow + ri}"

            if concept and concept not in seen:
                order.append(concept)
                seen.add(concept)
                c2line[concept] = line
                c2depth[concept] = depth

            for ci, period in pcols:
                vc = row[ci - 1] if ci - 1 < len(row) else None
                raw = vc.value if vc else None
                if raw is None or raw == "" or raw == "\u2014":
                    continue
                try:
                    value = float(raw)
                except (ValueError, TypeError):
                    continue
                if period.is_annual() and period.fiscal_year > max_fy:
                    max_fy = period.fiscal_year
                facts.append(FactRecord(
                    statement_type=stmt, concept=concept, line_label=line,
                    period=period, value=value,
                    source_file=fname, source_sheet=tab_name,
                    source_column=period.column_label, depth=depth,
                ))

        sheets.append(SheetData(
            source_file=fname, source_sheet=tab_name,
            statement_type=stmt, facts=facts, row_order=order,
            concept_to_line=c2line, concept_to_depth=c2depth,
        ))
        logger.info("  '%s' → %s | %d facts, %d concepts", tab_name, stmt, len(facts), len(order))

    wb.close()
    return WorkbookInfo(
        filename=fname, filepath=fp,
        is_10k=is_10k_filename(fname),
        latest_fy=max_fy, sheets=sheets,
    )


# ── load folder ────────────────────────────────────────────────────────────
def load_all_workbooks(folder: str | Path) -> list[WorkbookInfo]:
    folder = Path(folder)
    wbs: list[WorkbookInfo] = []
    for fp in sorted(folder.glob("*.xlsx")):
        try:
            wbs.append(load_workbook_data(fp))
        except Exception:
            logger.exception("Failed: %s", fp)
    logger.info("Loaded %d workbooks", len(wbs))
    normalize_fiscal_periods(wbs)
    return wbs


def pick_latest_10k(wbs: list[WorkbookInfo]) -> WorkbookInfo | None:
    """Select the latest 10-K workbook; fall back to latest-FY workbook."""
    tens = [w for w in wbs if w.is_10k]
    pool = tens if tens else wbs
    if not pool:
        return None
    return max(pool, key=lambda w: w.latest_fy)


# ── fiscal year normalization ──────────────────────────────────────────────

def _detect_quarter_offset(workbooks: list[WorkbookInfo]) -> bool:
    """Detect if quarter labels use calendar quarters instead of fiscal.

    For companies with non-December FY end, a bug in column-header
    generation causes some quarters to fall back to calendar labeling
    while others (whose month-end day matches the FY end day) keep
    correct fiscal labels.

    For September 30 FY end (e.g. FICO), the signature pattern in any
    sheet that has both FY and quarterly data is:

    ✓ FY(XX), 4Q(XX), 3Q(XX) — fiscal code path worked (Jun/Sep end-of-month OK)
    ✗ 1Q(XX) present but is really fiscal Q2 (calendar Q1 = Jan-Mar)
    ✗ 4Q(XX-1) present but is really fiscal Q1 (calendar Q4 = Oct-Dec)
    ✗ 2Q(XX) absent (calendar Q2 = Apr-Jun was correctly labeled 3Q by fiscal path)

    Returns True when this pattern is detected.
    """
    from collections import defaultdict

    for wb in workbooks:
        for sheet in wb.sheets:
            type_years: dict[str, set[int]] = defaultdict(set)
            for fact in sheet.facts:
                type_years[fact.period.period_type].add(fact.period.fiscal_year)

            fy_years = type_years.get("FY", set())
            for fy_yr in fy_years:
                has_4q = fy_yr in type_years.get("Q4", set())
                has_3q = fy_yr in type_years.get("Q3", set())
                has_2q = fy_yr in type_years.get("Q2", set())
                has_1q = fy_yr in type_years.get("Q1", set())
                has_4q_prev = (fy_yr - 1) in type_years.get("Q4", set())

                if has_4q and has_3q and not has_2q and has_4q_prev and has_1q:
                    logger.info(
                        "Detected non-December FY end in %s/%s: "
                        "FY%d has 4Q%d+3Q%d but also 4Q%d+1Q%d "
                        "(missing 1Q%d+2Q%d — calendar quarter labeling detected)",
                        wb.filename, sheet.source_sheet,
                        fy_yr, fy_yr, fy_yr, fy_yr - 1, fy_yr, fy_yr, fy_yr,
                    )
                    return True

            # Fallback: detect from cumulative-period sheets
            for anchor_type in ("9M", "6M"):
                for a_yr in type_years.get(anchor_type, set()):
                    q1_has = a_yr in type_years.get("Q1", set())
                    q4_prev = (a_yr - 1) in type_years.get("Q4", set())
                    q2_has = a_yr in type_years.get("Q2", set())
                    if q4_prev and q1_has and not q2_has:
                        logger.info(
                            "Detected non-December FY end from %s in %s/%s",
                            anchor_type, wb.filename, sheet.source_sheet,
                        )
                        return True

    return False


def normalize_fiscal_periods(workbooks: list[WorkbookInfo]) -> None:
    """Detect and fix calendar-based quarter labels for non-December FY end.

    When the XBRL-to-Excel conversion falls back to calendar quarters for
    some periods (due to end-of-month mismatch), the result is a mix of
    correct fiscal labels (3Q, 4Q) and mislabeled calendar labels (4Q of
    prior year = fiscal Q1, 1Q = fiscal Q2).

    This function detects the pattern and remaps on a per-sheet basis so
    that legitimate fiscal Q4 labels (in sheets with their own FY) are
    preserved while mislabeled ones are corrected.
    """
    if not _detect_quarter_offset(workbooks):
        return

    logger.info("Applying per-sheet fiscal quarter normalization")

    # Collect all FY years across all data to know valid fiscal years
    all_fy_years: set[int] = set()
    for wb in workbooks:
        for sheet in wb.sheets:
            for fact in sheet.facts:
                if fact.period.period_type == "FY":
                    all_fy_years.add(fact.period.fiscal_year)

    remapped = 0
    for wb in workbooks:
        for sheet in wb.sheets:
            remapped += _normalize_sheet(sheet, all_fy_years)

    logger.info("Fiscal period normalization: remapped %d facts", remapped)


def _normalize_sheet(sheet: SheetData, all_fy_years: set[int]) -> int:
    """Remap mislabeled calendar quarters in a single sheet.

    The rule for each fact:
      - ``1Q(XX)`` → ``2Q(XX)`` always  (calendar Q1 = fiscal Q2)
      - ``4Q(XX)`` → ``1Q(XX+1)`` ONLY when this sheet does NOT also
        contain ``FY(XX)`` or ``3Q(XX)`` (which would confirm ``4Q(XX)``
        is a legitimately-labeled fiscal Q4)
      - ``FY``, ``3Q``, ``6M``, ``9M`` — unchanged
    """
    from collections import defaultdict

    type_years: dict[str, set[int]] = defaultdict(set)
    for fact in sheet.facts:
        type_years[fact.period.period_type].add(fact.period.fiscal_year)

    # Identify years where 4Q is definitely the correct fiscal Q4
    # (sheet also has FY or 3Q for the same year)
    protected_q4_years: set[int] = set()
    for yr in type_years.get("Q4", set()):
        if yr in type_years.get("FY", set()) or yr in type_years.get("Q3", set()):
            protected_q4_years.add(yr)

    count = 0
    for fact in sheet.facts:
        pt = fact.period.period_type
        yr = fact.period.fiscal_year

        if pt == "Q1":
            # Calendar Q1 is fiscal Q2 (Jan-Mar → fiscal Q2 for Sep FY end)
            fact.period = Period("Q2", yr, fact.period.column_label)
            count += 1

        elif pt == "Q4" and yr not in protected_q4_years:
            # Calendar Q4 is fiscal Q1 of the NEXT fiscal year
            new_yr = yr + 1
            if new_yr in all_fy_years or any(
                y in all_fy_years for y in range(new_yr - 1, new_yr + 2)
            ):
                fact.period = Period("Q1", new_yr, fact.period.column_label)
                count += 1

    return count
