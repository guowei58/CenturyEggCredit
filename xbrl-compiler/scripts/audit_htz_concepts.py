"""Audit concept tags in every HTZ workbook — raw Excel Concept column, not loader inference."""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from workbook_loader import load_all_workbooks
from xbrl_periods import is_xbrl_tagged_concept, workbook_has_xbrl_tagged_facts

INPUT = Path(__file__).resolve().parents[2] / "scripts/.audit-truth-batch/work/HTZ/in"

STATEMENT_TABS = (
    "Income Statement",
    "Balance Sheet",
    "Cash Flow",
    "Cash Flow Statement",
    "Cash Flows",
)


def audit_raw_excel(filepath: Path) -> dict:
    """Read Concept column directly from xlsx — ground truth."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    pref: Counter[str] = Counter()
    samples: dict[str, list[str]] = {}
    sheets_found: list[str] = []

    for name in wb.sheetnames:
        if name not in STATEMENT_TABS and not any(
            k in name.lower() for k in ("income", "balance", "cash flow")
        ):
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(min_row=1, max_row=500, values_only=True))
        if not rows:
            continue
        header = rows[0]
        concept_col = None
        for i, h in enumerate(header):
            if h and str(h).strip().lower() == "concept":
                concept_col = i
                break
        if concept_col is None:
            continue
        sheets_found.append(name)
        for row in rows[1:]:
            if not row or concept_col >= len(row):
                continue
            c = row[concept_col]
            if c is None or str(c).strip() == "":
                continue
            cs = str(c).strip()
            if cs.startswith("us-gaap:"):
                bucket = "us-gaap"
            elif cs.startswith("html:"):
                bucket = "html"
            elif cs.startswith("htz:") or cs.startswith("htzz:"):
                bucket = "htz/htzz"
            elif is_xbrl_tagged_concept(cs):
                bucket = "other-xbrl"
            else:
                bucket = "other"
            pref[bucket] += 1
            if bucket not in samples:
                samples[bucket] = []
            if len(samples[bucket]) < 3:
                samples[bucket].append(cs)

    return {
        "file": filepath.name,
        "sheets": sheets_found,
        "concept_mix": dict(pref),
        "samples": samples,
        "has_usgaap": pref.get("us-gaap", 0) > 0,
        "has_html": pref.get("html", 0) > 0,
        "has_htz": pref.get("htz/htzz", 0) > 0,
    }


def audit_loader(filepath: Path) -> dict:
    from workbook_loader import load_workbook_data

    wi = load_workbook_data(filepath)
    pref = Counter()
    for sh in wi.sheets:
        for f in sh.facts:
            c = f.concept or ""
            if c.startswith("us-gaap:"):
                pref["us-gaap"] += 1
            elif c.startswith("html:"):
                pref["html"] += 1
            elif c.startswith("htz:") or c.startswith("htzz:"):
                pref["htz/htzz"] += 1
            elif is_xbrl_tagged_concept(c):
                pref["other-xbrl"] += 1
            else:
                pref["other"] += 1
    return {
        "loader_mix": dict(pref),
        "has_xbrl_tagged": workbook_has_xbrl_tagged_facts(wi),
    }


def main() -> None:
    if not INPUT.is_dir():
        print(f"Missing input dir: {INPUT}")
        return

    files = sorted(INPUT.glob("*.xlsx"))
    print(f"Scanning {len(files)} workbooks in {INPUT}\n")

    ten_q_usgaap = []
    ten_q_html_only = []
    ten_q_mixed = []

    for fp in files:
        raw = audit_raw_excel(fp)
        loader = audit_loader(fp)
        is_10q = "10-Q" in fp.name or "10_Q" in fp.name

        if not is_10q:
            continue

        tag = "HTML-ONLY"
        if raw["has_usgaap"] or raw["has_htz"]:
            if raw["has_html"]:
                tag = "MIXED"
                ten_q_mixed.append(fp.name)
            else:
                tag = "TAGGED-ONLY"
                ten_q_usgaap.append(fp.name)
        else:
            ten_q_html_only.append(fp.name)

        print(f"--- {fp.name[-55:]} [{tag}]")
        print(f"  RAW excel concepts: {raw['concept_mix']}")
        if raw["samples"].get("us-gaap"):
            print(f"    us-gaap samples: {raw['samples']['us-gaap']}")
        if raw["samples"].get("htz/htzz"):
            print(f"    htz samples: {raw['samples']['htz/htzz']}")
        if raw["samples"].get("html"):
            print(f"    html samples: {raw['samples']['html'][:2]}")
        print(f"  LOADER facts:       {loader['loader_mix']}  has_xbrl_tagged={loader['has_xbrl_tagged']}")
        if raw["concept_mix"] != loader["loader_mix"]:
            print("  *** MISMATCH raw excel vs loader ***")
        print()

    print("=" * 60)
    print(f"10-Q TAGGED (us-gaap/htz, no html): {len(ten_q_usgaap)}")
    print(f"10-Q MIXED (us-gaap/htz + html):     {len(ten_q_mixed)}")
    print(f"10-Q HTML-ONLY:                      {len(ten_q_html_only)}")
    if ten_q_usgaap:
        print("\nTagged 10-Q files:")
        for f in ten_q_usgaap:
            print(f"  {f}")
    if ten_q_mixed:
        print("\nMixed 10-Q files:")
        for f in ten_q_mixed:
            print(f"  {f}")


if __name__ == "__main__":
    main()
