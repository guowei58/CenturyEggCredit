"""Analyze HTZ live compile output — exact failure points with evidence."""
from __future__ import annotations

import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

WORK = Path(__file__).resolve().parents[2] / "scripts/.htz-live-compile"
OUT = WORK / "out"
REPORT = WORK / "failure-report.md"


def read_csv(name: str) -> list[dict]:
    p = OUT / name
    if not p.exists():
        return []
    with p.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def main() -> None:
    result_path = WORK / "compile-result.json"
    if not result_path.exists():
        print(f"Missing {result_path} — run scripts/run-htz-compile-latest.ts first")
        sys.exit(1)

    result = json.loads(result_path.read_text(encoding="utf-8"))
    audit = read_csv("source_audit_trail.csv")
    concept_map = read_csv("concept_to_row_map.csv")
    conflicts = read_csv("conflicts.csv")

    sums = [r for r in audit if r.get("source_method") == "summed_within_file"]
    sum_by_canon: Counter[str] = Counter()
    for r in sums:
        sum_by_canon[f"{r['statement_type']}|{r['canonical_row_id']}"] += 1

    # Bad concept map: html/htzz mapped to wrong us-gaap via fuzzy
    bad_maps = []
    for r in concept_map:
        raw = r.get("raw_concept", "")
        canon = r.get("canonical_row_id", "")
        notes = r.get("notes", "")
        if "depreciation" in raw.lower() and canon == "us-gaap:Revenues":
            bad_maps.append(r)
        if raw.startswith("html:") and "vehicle" in raw.lower() and "Interest" in canon:
            bad_maps.append(r)
        if "beginning-of-period" in raw and "CashCashEquivalents" in canon:
            bad_maps.append(r)
        if "Fuzzy label match" in notes and raw.startswith(("html:", "htzz:")):
            bad_maps.append(r)

    # htzz vs htz namespace: raw htzz concepts, master uses htz
    htzz_raw = [r for r in concept_map if r.get("raw_concept", "").startswith("htzz:")]
    htz_master = [r for r in concept_map if r.get("canonical_row_id", "").startswith("htz:")]

    # Missing revenues: headline 10-Q periods with no audit entry for us-gaap:Revenues
    rev_audit_periods = {
        r["output_period"]
        for r in audit
        if r.get("canonical_row_id") == "us-gaap:Revenues"
        and r.get("statement_type") == "income_statement"
    }
    headline_gaps = []
    for r in audit:
        if r.get("canonical_row_id") != "us-gaap:Revenues":
            continue
    # check IS export for empty Q1 on recent years
    try:
        import openpyxl

        wb = openpyxl.load_workbook(OUT / "consolidated_historical_financials.xlsx", read_only=True, data_only=True)
        ws = wb["IS_Quarterly"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1] == "Revenues":
                d = dict(zip(headers, row))
                for p in ["1Q23", "1Q24", "1Q25", "1Q26", "2Q24", "3Q24", "FY23", "FY24"]:
                    headline_gaps.append({"period": p, "value": d.get(p)})
                break
    except Exception as e:
        headline_gaps = [{"error": str(e)}]

    # Duplicate vehicle master rows
    master_rows = read_csv("master_presentation_rows.csv")
    vehicle_rows = [
        r
        for r in master_rows
        if r.get("display_label", "").strip().lower() in ("vehicle", "vehicles", "non-vehicle")
    ]

    wt = result.get("workbook_truth") or {}
    lines = [
        "# HTZ compiler failure report (latest live workbooks)",
        "",
        f"Input: `{WORK / 'in'}`",
        f"Master: `{result.get('master_file')}`",
        f"Mapped facts: {result.get('mapped_facts')} | Unresolved: {result.get('unresolved_count')} | Conflicts: {result.get('conflicts_count')}",
        f"Validation: {result.get('validation_passed')} passed / {result.get('validation_failed')} failed",
        f"Within-file SUM events: {len(sums)}",
        f"Workbook truth issues: {wt.get('issues_count', 0)}",
        "",
        "## Where the compiler goes wrong (ordered by impact)",
        "",
        "### 1. Stage 1 consolidation — `_resolve_multi_concept` sums unrelated tags on one row",
        "",
        f"**{len(sums)} cells** used `summed_within_file` (Priority C in `consolidator.py` when multiple",
        "raw concepts map to the same canonical row + period and no master/subtotal rule fires).",
        "",
        "Top affected canonical rows:",
    ]
    for key, n in sum_by_canon.most_common(12):
        lines.append(f"- `{key}` — {n} SUM events")

    lines.extend(["", "Examples (from `source_audit_trail.csv`):", ""])
    for r in sums[:8]:
        lines.append(
            f"- **{r['master_display_label']}** @ {r['output_period']} in `{r['source_file'][-45:]}`: "
            f"`{r['raw_concept']}` → {r['value']} ({r.get('derivation_formula', '')[:80]})"
        )

    lines.extend([
        "",
        "**Code path:** `consolidator.py` → `_resolve_multi_concept` → Priority C sum when",
        "Priority A (master concept present), A2 (preferred revenue tag), B (subtotal detect) all miss.",
        "",
        "### 2. Phase 2 concept mapping — fuzzy / ambiguous label match",
        "",
        f"**{len([r for r in concept_map if 'Fuzzy label match' in r.get('notes', '')])} fuzzy label matches**",
        f"**{len(bad_maps)} flagged bad/ risky maps** (depreciation→Revenues, vehicle→Interest, cash rollforward, fuzzy html/htzz→us-gaap).",
        "",
        "**Code path:** `master_presentation_builder.py` → `_scan_workbook_phase2` →",
        "`_labels_similar_for_merge` (IS/CF) or `_resolve_norm_label_match` with multiple `vehicle` candidates (BS returns None → Phase 4 new rows).",
        "",
        "Sample bad maps:",
    ])
    seen = set()
    for r in bad_maps[:15]:
        k = (r["raw_concept"], r["canonical_row_id"])
        if k in seen:
            continue
        seen.add(k)
        lines.append(f"- `{r['raw_concept']}` → `{r['canonical_row_id']}` ({r.get('notes', '')[:70]})")

    lines.extend([
        "",
        "### 3. Namespace split — `htzz:` (10-Q) vs `htz:` (10-K master)",
        "",
        f"- Raw `htzz:` concepts in map: **{len(htzz_raw)}**",
        f"- Master/canonical `htz:` rows: **{len(htz_master)}**",
        "- Local-name match fails across prefix; label match is used instead → collisions with `us-gaap:` rows.",
        "",
        "### 4. Master presentation bloat — duplicate segment labels",
        "",
        f"**{len(vehicle_rows)} master rows** with display label Vehicle/Vehicles/Non-vehicle (different canonical IDs).",
        "Phase 4 adds more `html:`/`htzz:` rows when BS label match is ambiguous.",
        "",
        "### 5. Compiled grid gaps (Revenues example)",
        "",
    ])
    for g in headline_gaps:
        lines.append(f"- {g}")

    lines.extend([
        "",
        "### 6. Validation failures",
        "",
        f"**{result.get('validation_failed')}** checks failed (see compiler validators — roll-forwards, totals, etc.).",
        "",
        "### 7. Workbook truth",
        "",
        f"**{wt.get('issues_count', 0)}** truth issues after truth loop ({wt.get('iterations', '?')} iterations).",
    ])
    if wt.get("issues"):
        for i in wt["issues"][:10]:
            lines.append(
                f"- [{i.get('issue')}] {i.get('statement_type')} `{i.get('canonical_row_id')}` "
                f"@ {i.get('period')} compiled={i.get('compiled_value')} workbook={i.get('workbook_value')}"
            )

    lines.extend([
        "",
        "## What is NOT the problem (on latest tagged workbooks)",
        "",
        "- 10-Q workbooks **are** XBRL-tagged (us-gaap + htzz + some html headers)",
        "- Not html-only label-only matching for all rows — local-name match works for us-gaap tags",
        "- Not stale audit-truth-batch input",
        "",
    ])

    REPORT.write_text("\n".join(lines), encoding="utf-8")
    print(REPORT)
    print("\n".join(lines[:80]))


if __name__ == "__main__":
    main()
