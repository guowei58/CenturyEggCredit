"""Export consolidated data to Excel workbook and CSV audit files."""
from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from period_parser import parse_period, sort_period_labels
from consolidator import ConsolidatedData, AuditEntry, Conflict
from master_presentation_builder import MasterRow, ConceptMapping
from row_mapper import UnresolvedRow

logger = logging.getLogger(__name__)

_PREFIX = {"income_statement": "IS", "balance_sheet": "BS", "cash_flow": "CF"}
_HF = Font(bold=True, size=10)
_HB = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_TH = Side(style="thin")
_BD = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)


# ── period helpers ─────────────────────────────────────────────────────────

def _q_periods(concepts: dict[str, dict]) -> list[str]:
    s: set[str] = set()
    for vals in concepts.values():
        for lbl in vals:
            p = parse_period(lbl)
            if p and (p.is_quarterly() or p.is_annual()):
                s.add(lbl)
    return sort_period_labels(list(s))


def _fy_periods(concepts: dict[str, dict]) -> list[str]:
    s: set[str] = set()
    for vals in concepts.values():
        for lbl in vals:
            p = parse_period(lbl)
            if p and p.is_annual():
                s.add(lbl)
    return sort_period_labels(list(s))


# ── row ordering ───────────────────────────────────────────────────────────

def _row_order(master_rows: list[MasterRow], st: str, concepts: dict) -> list[str]:
    """Return canonical_row_ids in master display_order, appending consolidated-only keys.

    **Every** master row is listed even when *concepts* has no entry for that id
    (no facts mapped yet, or values only in dropped sparse columns).  Previously
    we required ``canonical_row_id in concepts``, which hid balance-sheet lines
    that appear on the 10-K presentation but never received a consolidated cell.
    """
    ordered: list[str] = []
    seen: set[str] = set()
    for mr in sorted(
        (r for r in master_rows if r.statement_type == st), key=lambda r: r.display_order
    ):
        cid = mr.canonical_row_id
        if cid not in seen:
            ordered.append(cid)
            seen.add(cid)
    for c in sorted(concepts):
        if c not in seen:
            ordered.append(c)
            seen.add(c)
    return ordered


def _label_map(master_rows: list[MasterRow]) -> dict[tuple[str, str], str]:
    return {(r.statement_type, r.canonical_row_id): r.display_label for r in master_rows}


# ── sheet writing ──────────────────────────────────────────────────────────

def _ws_header(ws, headers: list[str]) -> None:
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.border = _HF, _HB, _BD
        c.alignment = Alignment(horizontal="center")


def _write_stmt_sheet(
    wb: Workbook, name: str,
    concepts: dict[str, dict[str, float | None]],
    periods: list[str],
    order: list[str],
    labels: dict[tuple[str, str], str],
    st: str,
) -> None:
    ws = wb.create_sheet(name)
    _ws_header(ws, ["Concept", "Line"] + periods)
    for ri, crid in enumerate(order, 2):
        vals = concepts.get(crid, {})
        disp = labels.get((st, crid), crid)
        ws.cell(row=ri, column=1, value=crid).border = _BD
        ws.cell(row=ri, column=2, value=disp).border = _BD
        for ci, p in enumerate(periods, 3):
            v = vals.get(p)
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = _BD
            if isinstance(v, (int, float)):
                c.number_format = "#,##0.00"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "C2"


# ── CSV writers ────────────────────────────────────────────────────────────

def _write_csv(path: Path, headers: list[str], rows: list[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)


# ── public entry point ─────────────────────────────────────────────────────

def export_all(
    data: ConsolidatedData,
    master_rows: list[MasterRow],
    concept_map: list[ConceptMapping],
    audit: list[AuditEntry],
    conflicts: list[Conflict],
    unresolved: list[UnresolvedRow],
    log_messages: list[str],
    output_dir: str | Path,
) -> dict[str, Any]:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    labels = _label_map(master_rows)
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    for st in ("income_statement", "balance_sheet", "cash_flow"):
        if st not in data:
            continue
        pfx = _PREFIX[st]
        concepts = data[st]
        order = _row_order(master_rows, st, concepts)

        qp = _q_periods(concepts)
        if qp:
            _write_stmt_sheet(wb, f"{pfx}_Quarterly", concepts, qp, order, labels, st)
        fp = _fy_periods(concepts)
        if fp:
            _write_stmt_sheet(wb, f"{pfx}_Annual", concepts, fp, order, labels, st)

    # Conflicts sheet
    ws = wb.create_sheet("Conflicts")
    ch = ["statement_type", "canonical_row_id", "period", "value", "source_file", "source_sheet", "source_column", "raw_concept", "resolution"]
    _ws_header(ws, ch)
    ri = 2
    for c in conflicts:
        for val, sf, ss, sc, rc in c.values:
            for ci, v in enumerate([c.statement_type, c.canonical_row_id, c.period, val, sf, ss, sc, rc, c.resolution], 1):
                ws.cell(row=ri, column=ci, value=v).border = _BD
            ri += 1

    # Unresolved sheet
    ws = wb.create_sheet("Unresolved")
    uh = ["source_file", "source_sheet", "statement_type", "line_label", "concept", "period_label", "value", "reason"]
    _ws_header(ws, uh)
    for ri2, u in enumerate(unresolved, 2):
        for ci, fld in enumerate(uh, 1):
            ws.cell(row=ri2, column=ci, value=getattr(u, fld, "")).border = _BD

    xlsx_path = out / "consolidated_historical_financials.xlsx"
    wb.save(str(xlsx_path))
    logger.info("Wrote %s", xlsx_path)

    # CSVs
    _write_csv(out / "master_presentation_rows.csv",
               ["statement_type", "canonical_row_id", "master_raw_concept", "display_label", "display_order", "depth"],
               [vars(r) for r in master_rows])

    _write_csv(out / "concept_to_row_map.csv",
               ["statement_type", "raw_concept", "canonical_row_id", "mapping_status", "notes"],
               [vars(m) for m in concept_map])

    _audit_cols = ["statement_type", "canonical_row_id", "master_display_label", "output_period",
                   "value", "source_file", "source_sheet", "source_column",
                   "raw_line_label", "raw_concept", "source_method", "derivation_formula"]
    _write_csv(out / "source_audit_trail.csv", _audit_cols, [vars(a) for a in audit])

    _write_csv(out / "conflicts.csv",
               ["statement_type", "canonical_row_id", "period", "value", "source_file",
                "source_sheet", "source_column", "raw_concept", "resolution"],
               [{"statement_type": c.statement_type, "canonical_row_id": c.canonical_row_id,
                 "period": c.period, "value": v, "source_file": sf, "source_sheet": ss,
                 "source_column": sc, "raw_concept": rc, "resolution": c.resolution}
                for c in conflicts for v, sf, ss, sc, rc in c.values])

    _write_csv(out / "unresolved_rows.csv",
               ["source_file", "source_sheet", "statement_type", "line_label", "concept", "period_label", "value", "reason"],
               [vars(u) for u in unresolved])

    with open(out / "processing_log.csv", "w", encoding="utf-8") as f:
        for msg in log_messages:
            f.write(msg + "\n")

    files = [str(xlsx_path)] + [
        str(out / n) for n in (
            "master_presentation_rows.csv", "concept_to_row_map.csv",
            "source_audit_trail.csv", "conflicts.csv", "unresolved_rows.csv", "processing_log.csv",
        )
    ]
    logger.info("Export complete: %d files", len(files))
    return {"files": files}
