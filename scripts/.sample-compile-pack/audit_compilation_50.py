#!/usr/bin/env python3
"""Full compilation audit for 50-ticker sample pack. Writes audit-report.json + AUDIT_SUMMARY.md."""
from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
STMT_SHEETS = {
    "IS_Quarterly": "income_statement",
    "IS_Annual": "income_statement",
    "BS_Quarterly": "balance_sheet",
    "BS_Annual": "balance_sheet",
    "CF_Quarterly": "cash_flow",
    "CF_Annual": "cash_flow",
}
STMT_TAB_NAMES = {"Income Statement", "Balance Sheet", "Cash Flow"}


def _is_num(v: Any) -> bool:
    if v is None or v == "" or v == "—":
        return False
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def analyze_excel(xlsx: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    out: dict[str, Any] = {}
    for name in wb.sheetnames:
        if name not in STMT_SHEETS and not name.endswith("_Quarterly") and not name.endswith("_Annual"):
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            out[name] = {"error": "no data rows"}
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        period_cols = [(i, h) for i, h in enumerate(headers[2:], start=2) if h]
        data_rows = rows[1:]
        total_rows = len(data_rows)
        fill: dict[str, str] = {}
        zero: list[str] = []
        sparse: list[str] = []
        for ci, plabel in period_cols:
            filled = sum(1 for r in data_rows if ci < len(r) and _is_num(r[ci]))
            fill[plabel] = f"{filled}/{total_rows}"
            if filled == 0:
                zero.append(plabel)
            elif total_rows and filled / total_rows < 0.20:
                sparse.append(plabel)
        out[name] = {
            "total_rows": total_rows,
            "period_columns": [p for _, p in period_cols],
            "n_period_columns": len(period_cols),
            "fully_empty_columns": zero,
            "sparse_columns_lt20pct": sparse,
            "column_fill": fill,
        }
    wb.close()
    return out


def parse_processing_log(log_path: Path) -> dict[str, Any]:
    text = log_path.read_text(encoding="utf-8", errors="replace")
    lines = text.splitlines()
    issues: list[dict[str, str]] = []
    loader_events: list[str] = []
    patterns = [
        (r"Dropping empty column '([^']+)'", "loader_drop_empty_column"),
        (r"'([^']+)': no header row", "loader_no_header"),
        (r"'([^']+)': all period columns filtered as sparse", "loader_all_sparse_filtered"),
        (r"'([^']+)': kept 0/(\d+) facts", "loader_xbrl_filter_zero_facts"),
        (r"'([^']+)': kept (\d+)/(\d+) facts \(XBRL-tagged", "loader_xbrl_filter_partial"),
        (r"Pipeline failed: (.+)$", "pipeline_failed"),
        (r"No facts extracted", "pipeline_no_facts"),
        (r"Workbook truth validation: (\d+) issue\(s\) remain", "workbook_truth_issues"),
        (r"\[(missing_value|extra_value|value_mismatch|missing_line|extra_line)\]", "workbook_truth_detail"),
    ]
    wt_issues: list[str] = []
    for line in lines:
        for pat, kind in patterns:
            m = re.search(pat, line)
            if m:
                loader_events.append(line.strip())
                issues.append({"kind": kind, "line": line.strip(), "match": m.groups()})
                if kind == "workbook_truth_detail":
                    wt_issues.append(line.strip())
                break
    compile_ok = "Step 8: Exporting" in text and "Pipeline failed" not in text
    m_loaded = re.search(r"Loaded (\d+) workbooks, (\d+) statement sheets, (\d+) facts", text)
    loaded_stats = None
    if m_loaded:
        loaded_stats = {
            "workbooks": int(m_loaded.group(1)),
            "sheets": int(m_loaded.group(2)),
            "facts": int(m_loaded.group(3)),
        }
    return {
        "compile_completed": compile_ok,
        "loaded_stats": loaded_stats,
        "loader_event_count": len(loader_events),
        "loader_events_sample": loader_events[:30],
        "classified_issues": issues,
        "workbook_truth_log_lines": wt_issues[:50],
    }


def audit_trail_periods(trail_path: Path) -> dict[str, set[str]]:
    by_stmt: dict[str, set[str]] = defaultdict(set)
    with trail_path.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            by_stmt[row["statement_type"]].add(row["output_period"])
    return dict(by_stmt)


def source_workbook_periods(in_dir: Path) -> dict[str, Any]:
    """Scan saved source workbooks for period column headers per statement tab."""
    per_file: list[dict[str, Any]] = []
    all_periods: dict[str, set[str]] = defaultdict(set)
    parse_failures: list[str] = []

    sys.path.insert(0, str(ROOT.parent.parent / "xbrl-compiler"))
    try:
        from period_parser import parse_workbook_period
        from workbook_loader import classify_tab, _find_header, _parse_column_period, _read_meta_period_map
        from openpyxl import load_workbook as xl_load
    except ImportError as e:
        return {"error": str(e)}

    for fp in sorted(in_dir.glob("*.xlsx")):
        if fp.name.startswith("~$"):
            continue
        entry: dict[str, Any] = {"file": fp.name, "sheets": {}}
        try:
            wb = xl_load(str(fp), read_only=True, data_only=True)
            meta = _read_meta_period_map(wb)
            is_10k = "10-K" in fp.name or "_10-K_" in fp.name
            for tab in wb.sheetnames:
                stmt = classify_tab(tab)
                if stmt is None:
                    continue
                ws = wb[tab]
                hdr = _find_header(ws, statement_type=stmt, is_10k=is_10k, meta_periods=meta.get(tab, {}))
                if hdr is None:
                    entry["sheets"][tab] = {"error": "no_header_row"}
                    continue
                hrow, hcells = hdr
                periods: list[str] = []
                for ci, val in hcells.items():
                    low = val.lower().strip()
                    if low in ("concept", "line", "depth"):
                        continue
                    p = _parse_column_period(
                        ci, val, statement_type=stmt, is_10k=is_10k, meta_periods=meta.get(tab, {})
                    )
                    if p:
                        periods.append(p.canonical)
                        all_periods[stmt].add(p.canonical)
                entry["sheets"][tab] = {"periods": sorted(set(periods)), "n_periods": len(set(periods))}
            wb.close()
        except Exception as ex:
            parse_failures.append(f"{fp.name}: {ex}")
        per_file.append(entry)

    return {
        "all_periods_by_statement": {k: sorted(v) for k, v in all_periods.items()},
        "files": per_file,
        "parse_failures": parse_failures,
    }


def compare_periods(
    excel: dict[str, Any],
    audit_periods: dict[str, set[str]],
    source_periods: dict[str, set[str]] | None,
) -> dict[str, Any]:
    cmp: dict[str, Any] = {}
    for sheet, stmt in STMT_SHEETS.items():
        ex = excel.get(sheet)
        if not ex or "period_columns" not in ex:
            continue
        excel_cols = set(ex["period_columns"])
        audit_cols = audit_periods.get(stmt, set())
        src_cols = (source_periods or {}).get(stmt, set())
        cmp[sheet] = {
            "excel_columns": sorted(excel_cols),
            "audit_trail_periods": sorted(audit_cols),
            "source_workbook_periods": sorted(src_cols) if src_cols else None,
            "in_audit_not_in_excel": sorted(audit_cols - excel_cols),
            "in_excel_not_in_audit": sorted(excel_cols - audit_cols),
            "in_source_not_in_excel": sorted(src_cols - excel_cols) if src_cols else None,
            "fully_empty_in_excel": ex.get("fully_empty_columns", []),
        }
    return cmp


def classify_root_causes(ticker_audit: dict[str, Any]) -> list[dict[str, str]]:
    causes: list[dict[str, str]] = []
    pl = ticker_audit.get("processing_log", {})
    for issue in pl.get("classified_issues", []):
        kind = issue["kind"]
        if kind == "loader_no_header":
            causes.append({
                "cause": "HEADER_PARSE_FAILURE",
                "detail": issue["line"],
                "code": "workbook_loader._find_header returns None",
                "fix_hint": "Extend period_parser / Meta sheet fallback for non-standard column headers.",
            })
        elif kind == "loader_xbrl_filter_zero_facts":
            causes.append({
                "cause": "XBRL_PERIOD_FILTER_DROPPED_ALL_FACTS",
                "detail": issue["line"],
                "code": "xbrl_periods.filter_facts_to_xbrl_periods",
                "fix_hint": "Sheet has numeric data but no us-gaap:/vendor QName concepts in period columns; review HTML-only face extract.",
            })
        elif kind == "loader_drop_empty_column":
            causes.append({
                "cause": "LOADER_DROPPED_ZERO_NUMERIC_COLUMN",
                "detail": issue["line"],
                "code": "workbook_loader._filter_sparse_columns",
                "fix_hint": "Column header parsed but every cell empty/em-dash; check header/period mapping not wrong column.",
            })
        elif kind == "pipeline_no_facts":
            causes.append({
                "cause": "ZERO_FACTS_AFTER_LOAD",
                "detail": issue["line"],
                "code": "main.py load_all_workbooks",
                "fix_hint": "All statement sheets failed header parse or XBRL filter; re-save workbooks from face extract.",
            })
        elif kind == "workbook_truth_detail":
            causes.append({
                "cause": "WORKBOOK_TRUTH_MISMATCH",
                "detail": issue["line"],
                "code": "workbook_truth.validate_compiled_against_workbooks",
                "fix_hint": "Headline filing cell differs from consolidated; check consolidation winner / derivation.",
            })

    excel = ticker_audit.get("excel", {})
    for sheet, data in excel.items():
        for col in data.get("fully_empty_columns", []):
            if col:
                causes.append({
                    "cause": "EXCEL_COLUMN_PRESENT_BUT_ALL_EMPTY",
                    "detail": f"{sheet} column {col} has header but 0/{data.get('total_rows', '?')} rows filled",
                    "code": "exporter._q_periods includes period if ANY row has value; empty cols = rows exist without values",
                    "fix_hint": "Expected for UI-padded quarters without derived values; or missing consolidation for that period.",
                })

    for sheet, cmp in ticker_audit.get("period_compare", {}).items():
        missing = cmp.get("in_source_not_in_excel") or []
        if missing:
            causes.append({
                "cause": "SOURCE_PERIODS_MISSING_FROM_COMPILED_EXCEL",
                "detail": f"{sheet}: {missing[:8]}{'...' if len(missing) > 8 else ''}",
                "code": "exporter._q_periods / display_min_fiscal_year / consolidation",
                "fix_hint": "Period existed in source workbook headers but no consolidated facts mapped; trace source_audit_trail.",
            })
        audit_missing = cmp.get("in_audit_not_in_excel") or []
        if audit_missing:
            causes.append({
                "cause": "AUDIT_HAS_PERIODS_NOT_EXPORTED_TO_EXCEL",
                "detail": f"{sheet}: {audit_missing[:8]}",
                "code": "exporter period selection vs audit trail",
                "fix_hint": "Audit trail has values for periods dropped from Excel column list — check _q_periods filter.",
            })

    return causes


def audit_ticker(ticker: str, pack: Path) -> dict[str, Any]:
    tdir = pack / ticker
    xlsx = tdir / "out" / "consolidated_historical_financials.xlsx"
    alt = pack / f"{ticker}_compiled_financials.xlsx"
    if not xlsx.exists() and alt.exists():
        xlsx = alt
    result: dict[str, Any] = {"ticker": ticker, "status": "missing"}

    in_dir = tdir / "in"
    if in_dir.exists():
        result["source_workbook_count"] = len(list(in_dir.glob("*.xlsx")))

    if not xlsx.exists():
        log_path = tdir / "out" / "processing_log.csv"
        if log_path.exists():
            result["processing_log"] = parse_processing_log(log_path)
            result["status"] = "compile_failed"
            result["root_causes"] = classify_root_causes(result)
        return result

    result["status"] = "compiled"
    result["excel_path"] = str(xlsx)
    result["excel"] = analyze_excel(xlsx)

    trail = tdir / "out" / "source_audit_trail.csv"
    audit_p: dict[str, set[str]] = {}
    if trail.exists():
        audit_p = audit_trail_periods(trail)
        result["audit_trail_period_counts"] = {k: len(v) for k, v in audit_p.items()}

    src_p = None
    if in_dir.exists():
        sw = source_workbook_periods(in_dir)
        result["source_workbook_scan"] = sw
        src_p = {k: set(v) for k, v in sw.get("all_periods_by_statement", {}).items()}

    result["period_compare"] = compare_periods(result["excel"], audit_p, src_p)

    for extra in ["processing_log.csv", "unresolved_rows.csv", "conflicts.csv"]:
        p = tdir / "out" / extra
        if p.exists():
            if extra.endswith(".csv"):
                with p.open(encoding="utf-8") as f:
                    rows = list(csv.DictReader(f))
                key = extra.replace(".csv", "")
                result[key] = {"row_count": len(rows), "sample": rows[:5]}

    log_path = tdir / "out" / "processing_log.csv"
    if log_path.exists():
        result["processing_log"] = parse_processing_log(log_path)

    result["root_causes"] = classify_root_causes(result)
    result["issue_counts"] = Counter(c["cause"] for c in result["root_causes"])
    return result


def main() -> None:
    manifest_path = ROOT / "manifest.json"
    tickers: list[str] = []
    seed = None
    if manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        tickers = manifest.get("tickers", [])
        seed = manifest.get("seed")

    if not tickers:
        tickers = [
            t.strip()
            for t in "PG,XOM,NVDA,BRK.B,MSFT,UNH,WMT,AAPL,AMZN,GOOGL,JNJ,HD,JPM,V,META,NKE,PFE,BA,SPGI,TMO,MCD,KO,AMD,CAT,PM,INTU,TXN,ORCL,MRK,COST,PEGA,NXST,BIO,EMR,ETN,FICO,PH,DECK,ITW,MIDD,BLCO,GIII,SCVL,GEN,PRGS,MGRC,CALX,CEVA,PDFS,MGPI".split(
                ","
            )
        ]

    # Include any ticker dirs with compiled output even if not in manifest yet
    for d in ROOT.iterdir():
        if d.is_dir() and (d / "out" / "consolidated_historical_financials.xlsx").exists():
            if d.name not in tickers:
                tickers.append(d.name)

    audits = [audit_ticker(t, ROOT) for t in tickers]
    compiled = [a for a in audits if a["status"] == "compiled"]
    failed = [a for a in audits if a["status"] != "compiled"]

    agg_causes: Counter[str] = Counter()
    for a in audits:
        agg_causes.update(a.get("issue_counts", {}))

    empty_col_tickers: list[str] = []
    for a in compiled:
        for sheet, data in a.get("excel", {}).items():
            if data.get("fully_empty_columns"):
                empty_col_tickers.append(f"{a['ticker']}/{sheet}")

    report = {
        "seed": seed,
        "tickers_requested": len(tickers),
        "compiled_count": len(compiled),
        "missing_or_failed_count": len(failed),
        "aggregate_root_cause_counts": dict(agg_causes),
        "tickers_with_empty_excel_columns": empty_col_tickers,
        "audits": audits,
        "generated_at": __import__("datetime").datetime.utcnow().isoformat() + "Z",
    }

    out_json = ROOT / "audit-report.json"
    out_json.write_text(json.dumps(report, indent=2, default=list), encoding="utf-8")

    lines = [
        "# 50-Company Compilation Audit",
        "",
        f"Seed: {seed}",
        f"Tickers in manifest: {len(tickers)}",
        f"Compiled successfully: {len(compiled)}",
        f"Missing / failed: {len(failed)}",
        "",
        "## Aggregate root-cause categories",
        "",
    ]
    for cause, cnt in agg_causes.most_common():
        lines.append(f"- **{cause}**: {cnt}")
    lines.extend(["", "## Tickers with fully empty Excel columns", ""])
    if empty_col_tickers:
        for x in empty_col_tickers[:40]:
            lines.append(f"- {x}")
        if len(empty_col_tickers) > 40:
            lines.append(f"- ... +{len(empty_col_tickers) - 40} more")
    else:
        lines.append("- None in compiled outputs scanned so far")

    lines.extend(["", "## Per-ticker status", ""])
    for a in audits:
        status = a["status"]
        causes = a.get("issue_counts", {})
        top = ", ".join(f"{k}={v}" for k, v in causes.most_common(3)) if causes else "clean"
        lines.append(f"- **{a['ticker']}**: {status} ({top})")

    (ROOT / "AUDIT_SUMMARY.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {out_json}")
    print(f"Compiled: {len(compiled)}/{len(tickers)}")


if __name__ == "__main__":
    main()
