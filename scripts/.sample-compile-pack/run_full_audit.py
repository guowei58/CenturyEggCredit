#!/usr/bin/env python3
"""Full audit of 50-ticker compile pack."""
from __future__ import annotations

import csv
import json
import re
import subprocess
import sys
from collections import Counter
from pathlib import Path

import openpyxl

pack = Path(__file__).resolve().parent
manifest = json.loads((pack / "manifest.json").read_text(encoding="utf-8"))
results = {r["ticker"]: r for r in manifest["results"]}
T50 = manifest["tickers"]
STMT_MAP = {
    "IS_Quarterly": "income_statement",
    "BS_Quarterly": "balance_sheet",
    "CF_Quarterly": "cash_flow",
}
COMPILER = pack.parent.parent / "xbrl-compiler"


def analyze_excel(ticker: str) -> dict | None:
    x = pack / ticker / "out" / "consolidated_historical_financials.xlsx"
    if not x.exists():
        x = pack / f"{ticker}_compiled_financials.xlsx"
    if not x.exists():
        return None
    wb = openpyxl.load_workbook(x, read_only=True, data_only=True)
    out: dict = {"sheets": {}, "empty_cols": [], "sparse_cols": []}
    for sn in wb.sheetnames:
        if not sn.endswith("_Quarterly"):
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        periods = [str(h) for h in rows[0][2:] if h]
        total = len(rows) - 1
        sheet = {"periods": periods, "fill": {}}
        for ci, pl in enumerate(periods, start=2):
            filled = sum(
                1
                for r in rows[1:]
                if ci < len(r) and r[ci] not in (None, "") and str(r[ci]) != "\u2014"
            )
            sheet["fill"][pl] = filled
            if filled == 0:
                out["empty_cols"].append(f"{sn}:{pl}")
            elif total and filled / total < 0.25:
                out["sparse_cols"].append({"sheet": sn, "period": pl, "fill": f"{filled}/{total}"})
        out["sheets"][sn] = sheet
    wb.close()
    return out


def compile_meta(ticker: str) -> dict | None:
    ind = pack / ticker / "in"
    if not ind.exists():
        return None
    n = len(list(ind.glob("*.xlsx")))
    if n < 2:
        return {"workbooks": n}
    outd = pack / ticker / "out-audit-scan"
    outd.mkdir(exist_ok=True)
    r = subprocess.run(
        [sys.executable, "main.py", "--input", str(ind), "--output", str(outd)],
        capture_output=True,
        text=True,
        cwd=str(COMPILER),
    )
    if r.returncode != 0:
        return {"compile_error": (r.stderr or r.stdout)[-800:]}
    return json.loads(r.stdout.strip().splitlines()[-1])


def loader_issues(ticker: str) -> dict:
    logp = pack / ticker / "out" / "processing_log.csv"
    if not logp.exists():
        return {}
    text = logp.read_text(encoding="utf-8", errors="replace")
    wt_re = re.compile(
        r"\[(missing_value|extra_value|value_mismatch|missing_line|extra_line)\]"
    )
    return {
        "no_header": len(re.findall(r"no header row", text)),
        "xbrl_zero": len(re.findall(r"kept 0/\d+ facts", text)),
        "drop_empty_col": len(re.findall(r"Dropping empty column", text)),
        "wt_issues": len(wt_re.findall(text)),
    }


def main() -> None:
    audits = []
    agg: Counter[str] = Counter()

    for t in T50:
        mr = results[t]
        entry: dict = {
            "ticker": t,
            "compileOk": mr.get("compileOk"),
            "filingsSaved": mr.get("filingsSaved"),
            "filingsFailed": mr.get("filingsFailed"),
            "pipeline_error": mr.get("error"),
        }
        ex = analyze_excel(t)
        if ex:
            entry["empty_cols"] = ex["empty_cols"]
            entry["sparse_cols"] = ex["sparse_cols"][:8]
            entry["sparse_count"] = len(ex["sparse_cols"])

        meta = compile_meta(t) if mr.get("compileOk") else None
        if meta and "xbrl_backed_periods_by_statement" in meta:
            entry["workbook_truth_issues"] = meta.get("workbook_truth", {}).get("issues_count", 0)
            entry["sheets_processed"] = meta.get("sheets_processed")
            entry["total_facts"] = meta.get("total_facts")
            entry["files_processed"] = meta.get("files_processed")
            missing_ytd: dict = {}
            extra_fy_in_q: dict = {}
            if ex:
                xbrl = meta["xbrl_backed_periods_by_statement"]
                for sheet, stmt in STMT_MAP.items():
                    if sheet not in ex["sheets"]:
                        continue
                    backed = set(xbrl.get(stmt, []))
                    cols = set(ex["sheets"][sheet]["periods"])
                    ytd_missing = sorted(
                        p for p in backed if p.startswith(("6M", "9M")) and p not in cols
                    )
                    fy_in_q = sorted(p for p in cols if p.startswith("FY"))
                    if ytd_missing:
                        missing_ytd[sheet] = ytd_missing
                    if fy_in_q:
                        extra_fy_in_q[sheet] = fy_in_q
            entry["missing_6m_9m_in_excel"] = missing_ytd
            entry["fy_columns_in_quarterly_sheet"] = extra_fy_in_q
            if missing_ytd:
                agg["MISSING_6M_9M_IN_EXCEL"] += 1
            if extra_fy_in_q:
                agg["FY_IN_QUARTERLY_SHEET"] += 1
            if entry.get("workbook_truth_issues", 0) > 0:
                agg["WORKBOOK_TRUTH_ISSUES"] += 1

        li = loader_issues(t)
        entry["loader"] = li
        if li.get("no_header", 0) > 0:
            agg["LOADER_NO_HEADER"] += 1
            entry["loader_no_header_count"] = li["no_header"]
        if ex and ex["empty_cols"]:
            agg["FULLY_EMPTY_EXCEL_COLUMNS"] += len(ex["empty_cols"])
        if not mr.get("compileOk"):
            agg["COMPILE_FAILED"] += 1
        if mr.get("filingsSaved", 0) < 20:
            agg["THIN_FILING_HISTORY"] += 1
        audits.append(entry)

    summary = {
        "seed": manifest["seed"],
        "tickers": len(T50),
        "compiled_ok": sum(1 for a in audits if a.get("compileOk")),
        "aggregate_issue_counts": dict(agg),
        "tickers_with_empty_excel_columns": [a["ticker"] for a in audits if a.get("empty_cols")],
        "tickers_with_workbook_truth_issues": [
            a["ticker"] for a in audits if a.get("workbook_truth_issues", 0) > 0
        ],
        "tickers_with_missing_ytd_columns": [
            a["ticker"] for a in audits if a.get("missing_6m_9m_in_excel")
        ],
        "tickers_with_loader_no_header": [
            a["ticker"] for a in audits if a.get("loader_no_header_count", 0) > 0
        ],
        "audits": audits,
    }
    (pack / "full-audit-report.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(f"compiled_ok {summary['compiled_ok']} / {summary['tickers']}")
    print("aggregate", json.dumps(summary["aggregate_issue_counts"], indent=2))
    print("empty cols tickers", summary["tickers_with_empty_excel_columns"])
    print("missing ytd tickers", len(summary["tickers_with_missing_ytd_columns"]))
    print("loader no header", summary["tickers_with_loader_no_header"])
    print("workbook truth issues", summary["tickers_with_workbook_truth_issues"])
    print("compile failed", [a["ticker"] for a in audits if not a.get("compileOk")])


if __name__ == "__main__":
    main()
