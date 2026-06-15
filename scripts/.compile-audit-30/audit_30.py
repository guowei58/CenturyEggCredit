#!/usr/bin/env python3
"""Evidence-based audit for 30-ticker compile batch. Writes report.json + AUDIT_REPORT.md."""
from __future__ import annotations

import csv
import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl

AUDIT_DIR = Path(__file__).resolve().parent
PACK = AUDIT_DIR.parent / ".sample-compile-pack"
COMPILER = PACK.parent.parent / "xbrl-compiler"
STMT_MAP = {
    "IS_Quarterly": "income_statement",
    "IS_YTD": "income_statement",
    "IS_Annual": "income_statement",
    "BS_Quarterly": "balance_sheet",
    "BS_YTD": "balance_sheet",
    "BS_Annual": "balance_sheet",
    "CF_Quarterly": "cash_flow",
    "CF_YTD": "cash_flow",
    "CF_Annual": "cash_flow",
}
SPARSE_THRESHOLD = 0.25


def _num(v: Any) -> bool:
    if v is None or v == "" or v == "\u2014":
        return False
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False


def analyze_excel(ticker: str) -> dict[str, Any]:
    x = PACK / ticker / "out" / "consolidated_historical_financials.xlsx"
    if not x.exists():
        x = PACK / f"{ticker}_compiled_financials.xlsx"
    if not x.exists():
        return {"missing": True}
    wb = openpyxl.load_workbook(x, read_only=True, data_only=True)
    out: dict[str, Any] = {"sheets": {}, "empty_cols": [], "sparse_cols": [], "has_ytd_sheets": False}
    for sn in wb.sheetnames:
        if sn.endswith("_YTD"):
            out["has_ytd_sheets"] = True
        if not any(sn.endswith(s) for s in ("_Quarterly", "_YTD", "_Annual")):
            continue
        rows = list(wb[sn].iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        periods = [str(h) for h in rows[0][2:] if h]
        total = len(rows) - 1
        sheet_info: dict[str, Any] = {"periods": periods, "fill": {}}
        for ci, pl in enumerate(periods, start=2):
            filled = sum(1 for r in rows[1:] if ci < len(r) and _num(r[ci]))
            sheet_info["fill"][pl] = f"{filled}/{total}"
            if filled == 0:
                out["empty_cols"].append(f"{sn}:{pl}")
            elif total and filled / total < SPARSE_THRESHOLD:
                out["sparse_cols"].append(
                    {"sheet": sn, "period": pl, "fill": f"{filled}/{total}", "pct": round(filled / total, 3)}
                )
        out["sheets"][sn] = sheet_info
    wb.close()
    return out


def compile_meta(ticker: str) -> dict[str, Any] | None:
    ind = PACK / ticker / "in"
    if not ind.exists() or len(list(ind.glob("*.xlsx"))) < 2:
        return None
    outd = PACK / ticker / "out-audit-30"
    outd.mkdir(exist_ok=True)
    r = subprocess.run(
        [sys.executable, "main.py", "--input", str(ind), "--output", str(outd)],
        capture_output=True,
        text=True,
        cwd=str(COMPILER),
    )
    line = r.stdout.strip().splitlines()[-1] if r.stdout.strip() else ""
    try:
        data = json.loads(line)
    except json.JSONDecodeError:
        return {"compile_error": (r.stderr or r.stdout)[-1000:], "exit_code": r.returncode}
    data["exit_code"] = r.returncode
    return data


def loader_stats(ticker: str) -> dict[str, Any]:
    sys.path.insert(0, str(COMPILER))
    from workbook_loader import load_workbook_data

    ind = PACK / ticker / "in"
    if not ind.exists():
        return {}
    total = empty = 0
    total_facts = 0
    no_header_by_sheet: Counter[str] = Counter()
    for fp in sorted(ind.glob("*.xlsx")):
        if fp.name.startswith("~$"):
            continue
        total += 1
        try:
            wb = load_workbook_data(fp)
        except Exception as ex:
            empty += 1
            continue
        if not wb.sheets:
            empty += 1
        else:
            total_facts += sum(len(s.facts) for s in wb.sheets)
    return {
        "workbooks": total,
        "workbooks_zero_sheets": empty,
        "total_facts_loaded": total_facts,
    }


def trace_sparse_to_source(ticker: str, sheet: str, period: str) -> dict[str, Any]:
    """Find best-matching source workbook column fill for a sparse compiled period."""
    sys.path.insert(0, str(COMPILER))
    from workbook_loader import (
        _find_header,
        _parse_column_period,
        _read_meta_period_map,
        classify_tab,
    )
    from openpyxl import load_workbook as xl_load

    stmt = STMT_MAP.get(sheet, "")
    if not stmt:
        return {"error": "unknown sheet"}
    ind = PACK / ticker / "in"
    best: dict[str, Any] | None = None
    for fp in sorted(ind.glob("*.xlsx")):
        try:
            wb = xl_load(str(fp), read_only=True, data_only=True)
            meta = _read_meta_period_map(wb)
            is_10k = "10-K" in fp.name
            for tab in wb.sheetnames:
                if classify_tab(tab) != stmt:
                    continue
                ws = wb[tab]
                hdr = _find_header(ws, statement_type=stmt, is_10k=is_10k, meta_periods=meta.get(tab, {}))
                if hdr is None:
                    continue
                hrow, hcells = hdr
                target_ci = None
                for ci, val in hcells.items():
                    p = _parse_column_period(
                        ci, val, statement_type=stmt, is_10k=is_10k, meta_periods=meta.get(tab, {})
                    )
                    if p and p.canonical == period:
                        target_ci = ci
                        break
                if target_ci is None:
                    continue
                filled = total = 0
                for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                    if not row or (not row[0] and (len(row) < 2 or not row[1])):
                        continue
                    total += 1
                    v = row[target_ci - 1] if target_ci - 1 < len(row) else None
                    if _num(v):
                        filled += 1
                hit = {
                    "source_file": fp.name,
                    "source_sheet": tab,
                    "source_fill": f"{filled}/{total}",
                    "header": hcells.get(target_ci, period),
                }
                if best is None or filled > int(best["source_fill"].split("/")[0]):
                    best = hit
            wb.close()
        except Exception:
            continue
    if best:
        compiled = PACK / ticker / "out" / "consolidated_historical_financials.xlsx"
        if compiled.exists():
            wb2 = openpyxl.load_workbook(compiled, read_only=True, data_only=True)
            if sheet in wb2.sheetnames:
                rows = list(wb2[sheet].iter_rows(values_only=True))
                hdr = rows[0]
                ci = next((i for i, h in enumerate(hdr) if str(h) == period), None)
                if ci is not None:
                    filled_c = sum(1 for r in rows[1:] if ci < len(r) and _num(r[ci]))
                    best["compiled_fill"] = f"{filled_c}/{len(rows)-1}"
            wb2.close()
        return best
    return {"source_match": None, "period": period}


def classify_issue(entry: dict[str, Any]) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []
    t = entry["ticker"]

    if not entry.get("compileOk"):
        err = entry.get("pipeline_error") or entry.get("compile_meta_error") or ""
        if "EBUSY" in err or entry.get("filingsFailed", 0) > 0:
            issues.append({
                "category": "DOWNLOAD_FILE_LOCK",
                "evidence": f"filingsFailed={entry.get('filingsFailed')}, saved={entry.get('filingsSaved')}",
                "root_cause": "fs.writeFile hit EBUSY/Permission denied — target path locked by another process (Excel, antivirus, concurrent audit script reading same in/ file).",
                "code": "scripts/compile-sample-pack.ts writeFileWithRetry",
                "fix": "Write to temp file then rename; skip re-download when in/ already has >=28 workbooks; avoid opening in/ files during bulk run.",
            })
        elif "invalid compiler JSON" in err or entry.get("compile_meta", {}).get("compile_error"):
            issues.append({
                "category": "COMPILER_EXIT_NONZERO",
                "evidence": err[:300],
                "root_cause": "Python main.py exited non-zero; stderr INFO logs may be captured as error even when partial output exists.",
                "code": "scripts/compile-sample-pack.ts runCompiler",
                "fix": "Parse JSON from stdout when consolidated_historical_financials.xlsx exists; surface workbook_truth.issues_count separately from log noise.",
            })
        else:
            issues.append({
                "category": "COMPILE_FAILED",
                "evidence": err[:300] if err else "compileOk=false",
                "root_cause": "See pipeline_error in manifest.",
                "code": "xbrl-compiler/main.py",
                "fix": "Inspect processing_log.csv for the ticker.",
            })

    meta = entry.get("compile_meta") or {}
    wt = (meta.get("workbook_truth") or {}).get("issues_count", 0)
    if wt:
        issues.append({
            "category": "WORKBOOK_TRUTH_MISMATCH",
            "evidence": f"issues_count={wt}",
            "root_cause": "Headline filing cell differs from consolidated grid after truth loop.",
            "code": "workbook_truth.validate_compiled_against_workbooks",
            "fix": "Inspect workbook_truth.issues in compile JSON; trace canonical_row_id + period in source_audit_trail.",
        })

    loader = entry.get("loader") or {}
    if loader.get("workbooks_zero_sheets", 0) > 0:
        issues.append({
            "category": "LOADER_SKIPPED_WORKBOOKS",
            "evidence": f"{loader['workbooks_zero_sheets']}/{loader['workbooks']} workbooks loaded 0 sheets",
            "root_cause": "workbook_loader._find_header returned None (unparseable period headers) or all period columns empty after _filter_sparse_columns.",
            "code": "workbook_loader.load_workbook_data",
            "fix": "Extend period_parser for failing header patterns; check Meta sheet period map fallback.",
        })

    ex = entry.get("excel") or {}
    if ex.get("empty_cols"):
        issues.append({
            "category": "EMPTY_EXCEL_COLUMN",
            "evidence": ", ".join(ex["empty_cols"][:5]),
            "root_cause": "Column header exported but no numeric values in any master row for that period.",
            "code": "exporter._write_stmt_sheet",
            "fix": "Omit periods with zero mapped values from export column list.",
        })

    for sp in (ex.get("sparse_cols") or [])[:3]:
        trace = sp.get("trace") or {}
        src = trace.get("source_fill")
        comp = trace.get("compiled_fill")
        if src and comp:
            src_n = src.split("/")
            if len(src_n) == 2 and int(src_n[0]) > 0:
                issues.append({
                    "category": "SPARSE_COLUMN_MATCHES_SOURCE",
                    "evidence": f"{sp['sheet']} {sp['period']}: compiled {comp}, source {src} ({trace.get('source_file', '')[:40]})",
                    "root_cause": "Master presentation has more rows than the filing column; not a consolidation drop.",
                    "code": "exporter._row_order + face extract row count",
                    "fix": "Display-only: flag low-fill columns; optional hide unmapped master rows per period.",
                })
        elif trace.get("source_match") is None:
            issues.append({
                "category": "SPARSE_NO_SOURCE_COLUMN",
                "evidence": f"{sp['sheet']} {sp['period']}: {sp['fill']}",
                "root_cause": "Period in compiled grid but no matching period column in any saved source workbook (derived or consolidated-only).",
                "code": "derivation_engine / consolidator",
                "fix": "Verify source_method in source_audit_trail for that period.",
            })

    if meta.get("xbrl_backed_periods_by_statement") and ex.get("sheets"):
        xbrl = meta["xbrl_backed_periods_by_statement"]
        for ytd_sheet, stmt in (("IS_YTD", "income_statement"), ("CF_YTD", "cash_flow")):
            backed_ytd = {p for p in xbrl.get(stmt, []) if p.startswith(("6M", "9M"))}
            if not backed_ytd:
                continue
            if not ex.get("has_ytd_sheets"):
                issues.append({
                    "category": "MISSING_YTD_SHEET",
                    "evidence": f"{len(backed_ytd)} YTD periods in model, no *_YTD Excel sheets",
                    "root_cause": "Exporter fix not applied or compile used stale out/ from before YTD sheets were added.",
                    "code": "exporter.export_all",
                    "fix": "Re-compile after exporter update; confirm IS_YTD/CF_YTD tabs present.",
                })
                break
            cols = set(ex["sheets"].get(ytd_sheet, {}).get("periods", []))
            missing = sorted(backed_ytd - cols)
            if missing:
                issues.append({
                    "category": "YTD_PERIODS_NOT_EXPORTED",
                    "evidence": f"{ytd_sheet} missing {missing[:6]}",
                    "root_cause": "xbrl_backed periods exist but _ytd_periods filter excluded them (display_min_fy or no values).",
                    "code": "exporter._ytd_periods",
                    "fix": "Check display_min_fiscal_year floor vs period years.",
                })

    if entry.get("filingsSaved", 0) < 20:
        issues.append({
            "category": "THIN_FILING_HISTORY",
            "evidence": f"filingsSaved={entry.get('filingsSaved')}",
            "root_cause": "SEC cache has fewer 10-K/10-Q since 2019 or bulk save skipped filings with empty face extract.",
            "code": "prepareBulkPresentedFilings / fetchFacePresentedStatements",
            "fix": "Expected for recent IPO/spinoffs; document shorter grid.",
        })

    return issues


def main() -> None:
    manifest_path = PACK / "manifest.json"
    if not manifest_path.exists():
        print("No manifest.json — run compile-sample-pack first")
        sys.exit(1)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    tickers = manifest.get("tickers", [])
    results = {r["ticker"]: r for r in manifest.get("results", [])}

    audits: list[dict[str, Any]] = []
    all_issues: list[dict[str, Any]] = []

    for i, t in enumerate(tickers):
        print(f"[{i+1}/{len(tickers)}] auditing {t}...")
        mr = results.get(t, {})
        entry: dict[str, Any] = {
            "ticker": t,
            "compileOk": mr.get("compileOk"),
            "filingsSaved": mr.get("filingsSaved"),
            "filingsFailed": mr.get("filingsFailed"),
            "pipeline_error": mr.get("error"),
        }
        entry["excel"] = analyze_excel(t)
        entry["loader"] = loader_stats(t)
        if mr.get("compileOk"):
            entry["compile_meta"] = compile_meta(t)
            if entry["compile_meta"] and entry["compile_meta"].get("compile_error"):
                entry["compile_meta_error"] = entry["compile_meta"]["compile_error"]

        for sp in entry["excel"].get("sparse_cols", [])[:5]:
            sp["trace"] = trace_sparse_to_source(t, sp["sheet"], sp["period"])

        entry["issues"] = classify_issue(entry)
        for iss in entry["issues"]:
            all_issues.append({"ticker": t, **iss})
        audits.append(entry)

    by_cat = Counter(i["category"] for i in all_issues)
    report = {
        "seed": manifest.get("seed"),
        "tickers": tickers,
        "compiled_ok": sum(1 for a in audits if a.get("compileOk")),
        "total": len(tickers),
        "issues_by_category": dict(by_cat),
        "audits": audits,
    }
    (AUDIT_DIR / "audit-report.json").write_text(json.dumps(report, indent=2), encoding="utf-8")

    lines = [
        "# 30-Company Compilation Audit (seed {})".format(manifest.get("seed")),
        "",
        f"Compiled OK: **{report['compiled_ok']}/{report['total']}**",
        "",
        "## Issues by category",
        "",
    ]
    for cat, n in by_cat.most_common():
        lines.append(f"- **{cat}**: {n}")
    lines.extend(["", "## Per-ticker issues", ""])
    for a in audits:
        if not a.get("issues"):
            lines.append(f"- **{a['ticker']}**: clean")
            continue
        lines.append(f"### {a['ticker']}")
        for iss in a["issues"]:
            lines.append(f"- **{iss['category']}** — {iss['evidence']}")
            lines.append(f"  - Root cause: {iss['root_cause']}")
            lines.append(f"  - Code: `{iss['code']}`")
            lines.append(f"  - Suggested fix: {iss['fix']}")
        lines.append("")

    (AUDIT_DIR / "AUDIT_REPORT.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"Done: {report['compiled_ok']}/{report['total']} compiled")
    print(f"Wrote {AUDIT_DIR / 'AUDIT_REPORT.md'}")


if __name__ == "__main__":
    main()
